from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from shutil import copyfile

import glob
import os
import time
import subprocess
import sys
import datetime
import imaplib
import email
import smtplib
import ssl

from lower_functions import Credenciales_Correo, SecureSend, f_Email_Preferencias, get_messages, get_attachments

def createGraphicsWorksheet():

    excelData = Workbook()

    graphicsWorksheet = excelData.active

    return graphicsWorksheet, excelData

def loadWorkbook(fileName):

    return load_workbook(fileName, data_only=True, read_only=True)
    
def getWorksheets():

    worksheets = []

    while "Exit" not in worksheets:

        print("Insert the worksheet name or type Exit to continue: ")

        worksheet = input()

        worksheets.append(worksheet)

    del worksheets[-1]

    return worksheets

def getParametersIndexes():

    parametersIndexes = {
        "Impactos": [1,1],
        "Tier": [4,1],
        "Valor Publicitario": [9,1],
        "Valor Informativo": [12,1],
        "Favorabilidad Mediatica": [15,1],
        "Quote de Vocero": [18,1],
        "Presencia de Vocero": [21,1],
        "Mencion de Marca": [24,1],
        "Tipo de Medio": [27,1],
        "Titular de Gestion": [30,1],
        "Tipo de Gestion": [33,1]
    }

    return parametersIndexes

def getIndexes(data, worksheets):

    indexes = {}

    for month in worksheets:

        index = 9

        while (True):

            if (data[month].cell(index,11).value == None):

                indexes[month] = index - 1

                break

            index += 1

    return indexes

def getCriteriaPerMonth(criteriaName, criteriaCols, data, indexes):

    graphicsInformation[criteriaName] = {}

    for sheet in indexes.keys():

        currentSheetData = data[sheet]

        graphicsInformation[criteriaName][sheet] = [currentSheetData[col + str(indexes[sheet])].value for col in criteriaCols ] 

    return graphicsInformation

def editCriteriaAllMonths(criteriaName):

    criteriaDict = graphicsInformation[criteriaName]

    dataPoints = len(criteriaDict[list(criteriaDict.keys())[0]])

    editList = [0 for i in range(dataPoints)]

    for key in criteriaDict.keys():

        index = 0
        for element in criteriaDict[key]:

            editList[index] += element

            index += 1

    graphicsInformation[criteriaName] = editList

def addCriteriaPerMonth(criteriaName):

    [col, row] = parametersIndexes[criteriaName]

    headers = tableHeaders[criteriaName]

    headerIndex = 0
    for header in headers:
        graphicsSheet.cell(row, col + headerIndex).value = header
        headerIndex += 1

    criteriaData = graphicsInformation[criteriaName]

    row += 1

    rowIndex = 0
    for month in criteriaData.keys():

        colIndex = 0
        graphicsSheet.cell(row + rowIndex, col + colIndex).value = month
        colIndex += 1
        for value in criteriaData[month]:

            graphicsSheet.cell(row + rowIndex, col + colIndex).value = value

            colIndex += 1

        rowIndex += 1
    
    colEnd = col + colIndex
    rowEnd = row + rowIndex
    [colStart, rowStart] = parametersIndexes[criteriaName]

    graphicsSheet.cell(rowEnd, colStart).value = "Total"

    for totalCol in range(colStart + 1, colEnd):

        totalValue = 0
        for totalRow in range(rowStart + 1, rowEnd):

            totalValue += float(graphicsSheet.cell(totalRow, totalCol).value)

        graphicsSheet.cell(rowEnd, totalCol).value = totalValue 

def addCriteriaAllMonths(criteriaName):

    [col, row] = parametersIndexes[criteriaName]

    headers = tableHeaders[criteriaName]

    columnNames = tableColumns[criteriaName]

    headerIndex = 0
    for header in headers:
        graphicsSheet.cell(row, col + headerIndex).value = header
        headerIndex += 1

    rowIndex = 1
    for column in columnNames:
        graphicsSheet.cell(row + rowIndex, col).value = column
        rowIndex += 1
    
    row += 1
    col += 1
    rowIndex = 0
    for value in graphicsInformation[criteriaName]:

        graphicsSheet.cell(row + rowIndex, col).value = value
        rowIndex += 1

    row += rowIndex  
    graphicsSheet.cell(row, col - 1).value = "Total"
    totalValue = 0
    for i in range(2, row):

        totalValue += float(graphicsSheet.cell(i, col).value)

    graphicsSheet.cell(row, col).value = totalValue

def getTipoMedio():

    tipoMedio = {}

    for month in indexes.keys():

        rowStart = 9
        col = 7

        for row in range(rowStart, indexes[month]):

            currentValue = excelData[month].cell(row, col).value 

            if (currentValue not in tipoMedio.keys()):

                tipoMedio[currentValue] = 1

            else:

                tipoMedio[currentValue] += 1
    
    return tipoMedio

def addTipoMedio(tipoMedio):

    [col, row] = parametersIndexes["Tipo de Medio"]

    graphicsSheet.cell(row, col).value = "Tipo de Medio"
    graphicsSheet.cell(row, col + 1).value = "Cantidad"

    row += 1

    rowIndex = 0
    for tipo in tipoMedio.keys():
        
        graphicsSheet.cell(row + rowIndex, col).value = tipo
        graphicsSheet.cell(row + rowIndex, col + 1).value = tipoMedio[tipo]

        rowIndex += 1

    row += rowIndex 
    graphicsSheet.cell(row, col).value = "Total"
    col += 1
    totalValue = 0
    for totalRow in range(2, row):

        totalValue += int(graphicsSheet.cell(totalRow, col).value)

    graphicsSheet.cell(row, col).value = totalValue

def getTitularGestion():

    titularGestion = {}

    for month in indexes.keys():

        rowStart = 9
        col = 5
        rowPrev = indexes[month] - 1

        for row in range(indexes[month] - 1, rowStart - 1, -1):

            currentValue = excelData[month].cell(row, col).value

            if (currentValue != None):

                if (currentValue not in titularGestion.keys()):

                    titularGestion[currentValue] = rowPrev - row

                else:

                    titularGestion[currentValue] += rowPrev - row

                rowPrev = row
    
    return titularGestion

def addTitularGestion(sortedTitularGestion):

    [col, row] = parametersIndexes["Titular de Gestion"]

    graphicsSheet.cell(row, col).value = "Titular de Gestion"
    graphicsSheet.cell(row, col + 1).value = "Cantidad de Publicaciones"

    row += 1

    rowIndex = 0
    for i in range(1, 16):

        graphicsSheet.cell(row + rowIndex, col).value = sortedTitularGestion[-i][0]
        graphicsSheet.cell(row + rowIndex, col + 1).value = sortedTitularGestion[-i][1]

        rowIndex += 1

def optionsTipoGestion():

    return ["Nota de Prensa", "Entrevista", "Comentario de vocero", "Mención del cliente", "Mención de Marca y Vocera", "Mención de Marca y Vocero"]

def getTipoGestion():

    tipoGestionCounter = {}
    tipoGestionPublicaciones = {}

    for month in indexes.keys():

        rowStart = 9
        col = 4
        rowPrev = indexes[month] - 1

        for row in range(indexes[month] - 1, rowStart - 1, -1):

            currentValue = excelData[month].cell(row, col).value

            if (currentValue != None):

                for option in optionsTipoGestion():

                    if option in currentValue:

                        currentValue = option

                if (currentValue not in tipoGestionCounter.keys()):

                    tipoGestionCounter[currentValue] = 1
                    tipoGestionPublicaciones[currentValue] = rowPrev - row  

                else:

                    tipoGestionCounter[currentValue] += 1
                    tipoGestionPublicaciones[currentValue] += rowPrev - row  

                rowPrev = row  
            
    return tipoGestionCounter, tipoGestionPublicaciones

def addTipoGestion(tipoGestionCounter, tipoGestionPublicaciones):

    [col, row] = parametersIndexes["Tipo de Gestion"]

    graphicsSheet.cell(row, col).value = "Tipo de Gestion"
    graphicsSheet.cell(row, col + 1).value = "Cantidad"

    graphicsSheet.cell(row, col + 3).value = "Tipo de Gestion"
    graphicsSheet.cell(row, col + 4).value = "Cantidad de Publicaciones"

    row += 1

    rowIndex = 0
    for tipo in tipoGestionCounter.keys():

        graphicsSheet.cell(row + rowIndex, col).value = tipo
        graphicsSheet.cell(row + rowIndex, col + 3).value = tipo

        graphicsSheet.cell(row + rowIndex, col+1).value = tipoGestionCounter[tipo]
        graphicsSheet.cell(row + rowIndex, col + 4).value = tipoGestionPublicaciones[tipo]

        rowIndex += 1

# -------- Download file

mail = f_Email_Preferencias()
msgs = get_messages(mail)

if (len(msgs) > 0):

    for email_id in msgs:

        preferencias = get_attachments(mail, msgs, email_id)

else:
    
    exit()

fileName = preferencias[0]

# -------- Initializing information

worksheets = getWorksheets()

parametersIndexes = getParametersIndexes()

excelData = load_workbook(fileName, data_only=True, read_only=True)

graphicsSheet, graphicsSpreadsheet = createGraphicsWorksheet()

indexes = getIndexes(excelData, worksheets)

# --------- Processing the data and creating the tables

graphicsInformation = {}

criteriaColsPerMonth = {
    "Impactos": ["K"],
    "Tier": ["O", "P", "Q"],
    "Valor Publicitario": ["L"],
    "Valor Informativo": ["M"],
    "Favorabilidad Mediatica": ["X", "Y", "Z"],
    "Quote de Vocero" : ["V","W"],
    "Presencia de Vocero": ["T","U"],
    "Mencion de Marca": ["R","S"]
}

tableHeaders = {
    "Impactos": ["Mes", "Impactos"],
    "Tier": ["Mes", "Tier 1", "Tier 2", "Tier 3"],
    "Valor Publicitario": ["Mes", "Valor Publicitario"],
    "Valor Informativo": ["Mes", "Valor Informativo"],
    "Favorabilidad Mediatica": ["Favorabilidad Mediatica", "Cantidad"],
    "Quote de Vocero": ["Quote de Vocero", "Cantidad"],
    "Presencia de Vocero": ["Presencia de Vocero", "Cantidad"],
    "Mencion de Marca": ["Mencion de Marca", "Cantidad"]
}

tableColumns = {
    "Favorabilidad Mediatica": ["Positiva", "Negativa", "Neutral"],
    "Quote de Vocero": ["Si", "No"],
    "Presencia de Vocero": ["Si", "No"],
    "Mencion de Marca": ["Si", "No"]
}

for criteria in criteriaColsPerMonth.keys():

    getCriteriaPerMonth(criteria, criteriaColsPerMonth[criteria], excelData, indexes)

editCriteriaAllMonths("Favorabilidad Mediatica")
editCriteriaAllMonths("Quote de Vocero")
editCriteriaAllMonths("Presencia de Vocero")
editCriteriaAllMonths("Mencion de Marca")

print(graphicsInformation)

addCriteriaPerMonth("Impactos")
addCriteriaPerMonth("Tier")
addCriteriaPerMonth("Valor Publicitario")
addCriteriaPerMonth("Valor Informativo")

addCriteriaAllMonths("Favorabilidad Mediatica")
addCriteriaAllMonths("Quote de Vocero")
addCriteriaAllMonths("Presencia de Vocero")
addCriteriaAllMonths("Mencion de Marca")

tipoMedio = getTipoMedio()
addTipoMedio(tipoMedio)

titularGestion = getTitularGestion()
sortedTitularGestion = sorted(titularGestion.items(), key = lambda kv: kv[1])
addTitularGestion(sortedTitularGestion)

tipoGestionCounter, tipoGestionPublicaciones = getTipoGestion()
addTipoGestion(tipoGestionCounter, tipoGestionPublicaciones)


# --------- Saving files

graphicsSpreadsheet.save("Finished.xlsx")

# --------- Send Email

cred = Credenciales_Correo()

message = MIMEMultipart()

message["From"] = cred["sender_email"]
message["To"] = cred["receiver_email"]

message["Subject"] = "Graphics"

with open("Finished.xlsx", "rb") as attachment:

    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())

encoders.encode_base64(part)
message.attach(part)
txt = message.as_string()

SecureSend(txt, cred)

# ----------- remove files

os.remove(fileName)
os.remove("Finished.xlsx")



